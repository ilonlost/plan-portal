from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook

from app.services.import_service import ExcelImportService


def workbook_bytes(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_ohl_template_preserves_date_and_converts_pieces_to_full_boxes() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ОХЛ"
    sheet.append([None, None, None, None, None, None, "План производства"])
    sheet.append(["Сегмент", "Статус", "СГ", "Маркировка", "SAP-код", "Наименование", "12 авг"])
    sheet.append(["ОХЛ", "АЗ", 10, "Да", 101, "Курица 150г*4 (0,6кг)", Decimal("5")])
    ref = workbook.create_sheet("Справочник ФК")
    ref.append(["Код", "Наименование", "Линия", "Категория", "Скорость"])
    ref.append([101, "Курица", "Миквак", "Кулинария", 240])

    preview = ExcelImportService().parse(workbook_bytes(workbook), "ОХЛ 2026.xlsx")
    row = preview.rows[0]
    assert preview.template_type == "ohl_daily"
    assert row.requested_date == row.due_date == date(2026, 8, 12)
    assert row.source_quantity == Decimal("5")
    assert row.quantity_kg == Decimal("1.200")
    assert row.box_count == Decimal("2")
    assert row.warnings


def test_reference_template_reads_speed_batch_and_restriction() -> None:
    workbook = Workbook()
    workbook.active.title = "ПЦ"
    workbook.create_sheet("КЦ 1")
    workbook.create_sheet("КЦ 2")
    ref = workbook.create_sheet("Справочник")
    ref.append(["Код", "Наименование", "Состояние", "Линия", "Категория", "Кратко", "Код", "Скорость", "Код", "Замес", "Тип", "Расчёт", "Статус", "Чел", "Чел-ч", "Мин заказ", "Ограничение"])
    ref.append([101, "Круассан 70г*70 (4,9кг)", "ЗАМ", "Слойка", "ПЦ", "Круассан", 101, 462, 101, Decimal("273.5686"), "Дежа", None, "Активный", 10, None, Decimal("820.7058"), "Расстойка"])

    preview = ExcelImportService().parse(workbook_bytes(workbook), "План производства 12.08.2026.xlsm")
    row = preview.rows[0]
    assert preview.template_type == "production_reference"
    assert row.speed_kg_hour == Decimal("462")
    assert row.batch_quantum_kg == Decimal("273.5686")
    assert row.units_per_box == Decimal("70.000")
    assert row.restrictions == "Расстойка"


def test_sandwich_advance_marking_moves_production_one_day_back() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ОХЛ"
    sheet.append([None, None, None, None, None, None, "План производства"])
    sheet.append(["Сегмент", "Статус", "СГ", "Маркировка", "SAP-код", "Наименование", "12 авг"])
    sheet.append(["ОХЛ", "АЗ", 10, "Да", 101, "Сэндвич 150г*4 (0,6кг)", Decimal("4")])
    ref = workbook.create_sheet("Справочник ФК")
    ref.append(["Код", "Наименование", "Линия", "Категория", "Скорость"])
    ref.append([101, "Сэндвич", "Сэндвичи", "Кулинария", 240])

    row = ExcelImportService().parse(workbook_bytes(workbook), "ОХЛ 2026.xlsx").rows[0]
    assert row.advance_marking is True
    assert row.requested_date == date(2026, 8, 11)
    assert row.marking_date == date(2026, 8, 12)
