from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.config import settings


class ExcelExportService:
    HEADERS = [
        "№", "Дата", "Смена", "Линия", "Тип записи", "Продукт", "Артикул",
        "Количество источника", "Ед. источника", "Задание, шт.", "Задание, кг", "Коробов", "Квантов",
        "Время, ч", "Загрузка смены, %", "Причина", "Статус", "Источник", "Предупреждения",
    ]

    def __init__(self) -> None:
        self.file_name = "План производства.xlsx"
        self.media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def build(self, items: list[dict]) -> bytes:
        template = Path(settings.plan_export_template) if settings.plan_export_template else None
        if template and template.is_file() and template.suffix.lower() == ".xlsm":
            return self._build_from_template(template, items)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ПЦ"
        self._write_sheet(sheet, [item for item in items if item.get("workshop_code") == "PC"])
        kitchen = workbook.create_sheet("КЦ")
        self._write_sheet(kitchen, [item for item in items if item.get("workshop_code") != "PC"])
        all_items = workbook.create_sheet("План выгрузки")
        self._write_sheet(all_items, items)
        self.file_name = "План производства.xlsx"
        self.media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def _build_from_template(self, template: Path, items: list[dict]) -> bytes:
        workbook = load_workbook(template, keep_vba=True)
        sheet = workbook["План выгрузки"] if "План выгрузки" in workbook.sheetnames else workbook.create_sheet("План выгрузки")
        sheet.delete_rows(1, sheet.max_row)
        self._write_sheet(sheet, items)
        self.file_name = template.name
        self.media_type = "application/vnd.ms-excel.sheet.macroEnabled.12"
        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def _write_sheet(self, sheet, items: list[dict]) -> None:
        sheet.append(self.HEADERS)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="C8102E")
            cell.alignment = Alignment(horizontal="center")
        for item in items:
            sheet.append([
                item.get("sequence"), item.get("production_date"), "Ночь" if item.get("shift") == "night" else "День",
                item.get("line_name") or "—", item.get("schedule_kind"), item.get("product_name"), item.get("sku"),
                self._float(item.get("source_quantity")), item.get("source_unit"),
                self._float(item.get("quantity_units")), self._float(item.get("quantity_kg") or item.get("quantity")), self._float(item.get("box_count")),
                self._float(item.get("batch_count")), float(item.get("required_hours", 0)),
                float(item.get("load_percent", 0)), item.get("reason"), item.get("status"),
                item.get("source"), "; ".join(item.get("warnings", [])),
            ])
            status = item.get("status")
            color = "FCE8EC" if status in {"conflict", "unscheduled"} else "FFF5DF" if status == "warning" else "EAF8F1"
            for cell in sheet[sheet.max_row]:
                cell.fill = PatternFill("solid", fgColor=color)
        widths = [8, 14, 10, 24, 16, 30, 16, 20, 14, 16, 16, 12, 12, 14, 20, 30, 16, 14, 44]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    @staticmethod
    def _float(value):
        return float(value) if value is not None else None
