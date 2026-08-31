from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_CEILING
from io import BytesIO

from openpyxl import load_workbook

from app.schemas.common import ImportPreview, ImportRow


class ExcelImportService:
    """Parser for the real FK planning workbooks plus a generic tabular fallback."""

    COLUMN_ALIASES = {
        "sku": {"артикул", "sku", "код продукции", "код"},
        "product_name": {"продукт", "наименование", "наименование продукции", "product"},
        "quantity": {"количество", "потребность", "объем", "объём", "quantity"},
        "requested_date": {"желаемая дата", "дата производства", "requested date", "дата"},
        "due_date": {"крайняя дата", "срок", "due date", "дедлайн"},
        "priority": {"приоритет", "priority"},
        "customer": {"заказчик", "клиент", "customer"},
    }
    REQUIRED = {"sku", "quantity", "requested_date", "due_date"}

    def parse(self, content: bytes, file_name: str) -> ImportPreview:
        workbook = load_workbook(
            BytesIO(content), read_only=True, data_only=True,
        )
        try:
            if "ОХЛ" in workbook.sheetnames:
                return self._parse_ohl_daily(workbook, file_name)
            if "План ЗАМ+Напитки" in workbook.sheetnames:
                return self._parse_quarter_weekly(workbook, file_name)
            if "Справочник" in workbook.sheetnames and any(name in workbook.sheetnames for name in ("ПЦ", "КЦ 1", "КЦ 2")):
                return self._parse_production_reference(workbook, file_name)
            if "Справочник ФК" in workbook.sheetnames:
                return self._parse_capacity_reference(workbook, file_name)
            if any(name.strip() == "Рецептура 50" for name in workbook.sheetnames) and "План_пекарня (2)" in workbook.sheetnames:
                return self._parse_legacy_reference(workbook, file_name)
            return self._parse_generic(workbook, file_name)
        finally:
            workbook.close()

    def _parse_ohl_daily(self, workbook, file_name: str) -> ImportPreview:
        sheet = workbook["ОХЛ"]
        reference = self._fk_reference(workbook)
        year = self._year_from_workbook(sheet, file_name)
        headers = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        date_columns = [
            (column, parsed)
            for column, value in enumerate(headers)
            if (parsed := self._header_date(value, year))
        ]
        if not date_columns:
            raise ValueError("На листе ОХЛ не найдены календарные даты во второй строке")

        rows: list[ImportRow] = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            sku = self._sku(values[4] if len(values) > 4 else None)
            name = str((values[5] if len(values) > 5 else None) or "").strip()
            if not sku or not name:
                continue
            packaging = self.parse_packaging(name)
            ref = reference.get(sku, {})
            advance_value = str((values[3] if len(values) > 3 else None) or "").strip().lower()
            line_name = str(ref.get("line") or "").lower()
            product_text = name.lower().replace("ё", "е")
            is_advance_product = any(token in line_name or token in product_text for token in ("сэндвич", "сендвич", "бургер"))
            advance_marking = is_advance_product and advance_value not in {"", "нет", "0", "false", "none"}
            for column, source_date in date_columns:
                source_value = self._number(values[column] if column < len(values) else None)
                if source_value is None or source_value <= 0:
                    continue
                errors: list[str] = []
                warnings: list[str] = []
                unit_weight = packaging["unit_weight_kg"]
                units_per_box = packaging["units_per_box"]
                if unit_weight is None:
                    errors.append("Не удалось определить вес единицы из наименования")
                    quantity_kg = None
                    boxes = None
                else:
                    planned_pieces = source_value
                    boxes = None
                    if units_per_box and units_per_box > 0:
                        boxes = (source_value / units_per_box).quantize(Decimal("0.001"))
                        rounded_boxes = (source_value / units_per_box).to_integral_value(rounding=ROUND_CEILING)
                        planned_pieces = rounded_boxes * units_per_box
                        if planned_pieces != source_value:
                            warnings.append(
                                f"{self._display_number(source_value)} шт. не делятся на короб "
                                f"{self._display_number(units_per_box)} шт.; задание округлено до "
                                f"{self._display_number(planned_pieces)} шт. ({self._display_number(rounded_boxes)} кор.)"
                            )
                            boxes = rounded_boxes
                    quantity_kg = (planned_pieces * unit_weight).quantize(Decimal("0.001"))
                rows.append(ImportRow(
                    row_number=row_number, sku=sku, product_name=name,
                    quantity=quantity_kg, source_quantity=source_value, source_unit="шт",
                    quantity_kg=quantity_kg, unit_weight_kg=unit_weight,
                    units_per_box=units_per_box, box_weight_kg=packaging["box_weight_kg"], box_count=boxes,
                    requested_date=source_date - timedelta(days=1) if advance_marking else source_date,
                    due_date=source_date - timedelta(days=1) if advance_marking else source_date,
                    source_plan_date=source_date, marking_date=source_date,
                    advance_marking=advance_marking,
                    production_week=(source_date - timedelta(days=1) if advance_marking else source_date).isocalendar().week,
                    exact_date=True,
                    line_hint=ref.get("line"), speed_kg_hour=ref.get("speed"), category=ref.get("category"),
                    valid=not errors, errors=errors, warnings=warnings,
                ))
        return self._preview(
            file_name, "ohl_daily_v1", "ohl_daily", "ОХЛ", rows,
            ["Даты ОХЛ зафиксированы: алгоритм не переносит объём на соседние дни.",
             "Для сэндвичей и бургеров с признаком авансовой маркировки ДП ставится на один день раньше ДМ.",
             "Значения источника трактуются как штуки и переводятся в кг по весу из наименования.",
             "Задание округляется вверх до целого короба; отклонение показывается предупреждением."],
        )

    def _parse_quarter_weekly(self, workbook, file_name: str) -> ImportPreview:
        sheet = workbook["План ЗАМ+Напитки"]
        reference = self._fk_reference(workbook)
        week_columns: list[tuple[int, int]] = []
        headers = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        for column, raw_value in enumerate(headers):
            value = str(raw_value or "")
            match = re.search(r"(\d{1,2})\s*(?:w|нед)", value, re.IGNORECASE)
            if match:
                week_columns.append((column, int(match.group(1))))
        if not week_columns:
            raise ValueError("На листе «План ЗАМ+Напитки» не найдены недельные колонки")

        year = self._year_from_workbook(sheet, file_name)
        rows: list[ImportRow] = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            sku = self._sku(values[1] if len(values) > 1 else None)
            name = str((values[2] if len(values) > 2 else None) or "").strip()
            if not sku or not name:
                continue
            packaging = self.parse_packaging(name)
            ref = reference.get(sku, {})
            for column, week in week_columns:
                quantity = self._number(values[column] if column < len(values) else None)
                if quantity is None or quantity <= 0:
                    continue
                monday = date.fromisocalendar(year, week, 1)
                rows.append(ImportRow(
                    row_number=row_number, sku=sku, product_name=name, quantity=quantity,
                    source_quantity=quantity, source_unit="кг", quantity_kg=quantity,
                    unit_weight_kg=packaging["unit_weight_kg"], units_per_box=packaging["units_per_box"],
                    box_weight_kg=packaging["box_weight_kg"], requested_date=monday,
                    due_date=date.fromisocalendar(year, week, 7), production_week=week, exact_date=False,
                    line_hint=ref.get("line"), speed_kg_hour=ref.get("speed"), category=ref.get("category"),
                ))
        return self._preview(
            file_name, "quarter_weekly_v1", "quarter_weekly", "План ЗАМ+Напитки", rows,
            ["Объёмы источника трактуются как кг.",
             "Каждая потребность распределяется только внутри своей ISO-недели.",
             "При расчёте применяется квант замеса из актуального справочника ПЦ/КЦ."],
        )

    def _parse_production_reference(self, workbook, file_name: str) -> ImportPreview:
        sheet = workbook["Справочник"]
        rows: list[ImportRow] = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            value = lambda index: values[index] if index < len(values) else None
            sku = self._sku(value(0))
            name = str(value(1) or "").strip()
            if not sku or not name:
                continue
            line = str(value(3) or "").strip() or None
            speed = self._number(value(7))
            errors: list[str] = []
            if not line:
                errors.append("Не указана производственная линия")
            if speed is None or speed <= 0:
                errors.append("Не указана положительная скорость линии")
            packaging = self.parse_packaging(name)
            rows.append(ImportRow(
                row_number=row_number, sku=sku, product_name=name, source_unit="кг",
                unit_weight_kg=packaging["unit_weight_kg"], units_per_box=packaging["units_per_box"],
                box_weight_kg=packaging["box_weight_kg"], line_hint=line, speed_kg_hour=speed,
                batch_quantum_kg=self._number(value(9)), min_order_kg=self._number(value(15)),
                capacity_type=str(value(10) or "").strip() or None,
                restrictions=str(value(16) or "").strip() or None,
                state=str(value(2) or "").strip() or None,
                category=str(value(4) or "").strip() or None,
                short_name=str(value(5) or "").strip() or None,
                reference_source=file_name,
                valid=not errors, errors=errors,
            ))
        return self._preview(
            file_name, "pc_kc_reference_v1", "production_reference", "Справочник", rows,
            ["Импорт обновляет продукцию, линии, скорости, кванты замеса, минимальные партии и ограничения.",
             "Строки ПЦ/КЦ используются как эталон структуры; повреждённые формулы #REF! не импортируются.",
             "После справочника можно загружать недельный или дневной план спроса."],
        )

    def _parse_capacity_reference(self, workbook, file_name: str) -> ImportPreview:
        """Import the current line-capacity catalogue supplied by production."""
        sheet = workbook["Справочник ФК"]
        rows: list[ImportRow] = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            value = lambda index: values[index] if index < len(values) else None
            sku = self._sku(value(0))
            name = str(value(1) or "").strip()
            if not sku or not name:
                continue
            line = str(value(3) or "").strip() or None
            speed = self._number(value(5))
            errors: list[str] = []
            if not line:
                errors.append("Не указана производственная линия")
            if speed is None or speed <= 0:
                errors.append("Не указана положительная скорость линии")
            hours = self._number(value(6))
            status = str(value(8) or "").strip() or None
            note = str(value(13) or "").strip() or None
            restrictions = "; ".join(part for part in [note, f"Статус: {status}" if status else None, f"Часов в смену: {hours}" if hours is not None else None] if part) or None
            rows.append(ImportRow(
                row_number=row_number, sku=sku, product_name=name, source_unit="кг",
                unit_weight_kg=self._number(value(2)), line_hint=line, category=str(value(4) or "").strip() or None,
                speed_kg_hour=speed, available_hours=hours, line_status=status, restrictions=restrictions,
                reference_source=file_name, valid=not errors, errors=errors,
            ))
        return self._preview(
            file_name, "line_capacity_reference_v1", "capacity_reference", "Справочник ФК", rows,
            ["Импорт обновляет актуальные скорости линий, статусы и технологические ограничения из файла мощности линий."],
        )

    def _parse_legacy_reference(self, workbook, file_name: str) -> ImportPreview:
        """Read the old workbook as a comparison/reference source.

        The old plan stores batch and daily capacity in pieces. Recipe detail is
        summarized as the number of distinct components per finished SKU; the raw
        workbook remains the source of truth for future material-requirement work.
        """
        recipe_name = next(name for name in workbook.sheetnames if name.strip() == "Рецептура 50")
        recipe_sheet = workbook[recipe_name]
        component_counts: dict[str, set[str]] = {}
        for values in recipe_sheet.iter_rows(min_row=2, values_only=True):
            sku = self._sku(values[1] if len(values) > 1 else None)
            material = self._sku(values[3] if len(values) > 3 else None)
            if sku and material:
                component_counts.setdefault(sku, set()).add(material)

        rows_by_sku: dict[str, ImportRow] = {}
        plans = (
            ("План_пекарня (2)", 1, 2, 3, 4, 6),
            ("План_кулинария", 2, 3, 4, None, 6),
        )
        for sheet_name, sku_col, name_col, quantum_col, unit_col, capacity_col in plans:
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            for row_number, values in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                value = lambda index: values[index] if index < len(values) else None
                sku = self._sku(value(sku_col))
                name = str(value(name_col) or "").strip()
                if not sku or not name:
                    continue
                group = str(value(0) or "").strip() or None
                packaging = self.parse_packaging(name)
                rows_by_sku[sku] = ImportRow(
                    row_number=row_number, sku=sku, product_name=name, source_unit="шт",
                    unit_weight_kg=packaging["unit_weight_kg"], units_per_box=packaging["units_per_box"],
                    box_weight_kg=packaging["box_weight_kg"], line_hint=group,
                    legacy_quantum_units=self._number(value(quantum_col)),
                    legacy_daily_capacity_units=self._number(value(capacity_col)),
                    legacy_capacity_unit=str(value(unit_col) if unit_col is not None else "шт").strip() or "шт",
                    recipe_component_count=len(component_counts.get(sku, set())),
                    reference_source=file_name,
                )
        for sku, materials in component_counts.items():
            if sku not in rows_by_sku:
                rows_by_sku[sku] = ImportRow(
                    row_number=1, sku=sku, product_name=sku, source_unit="шт",
                    recipe_component_count=len(materials), reference_source=file_name,
                )
        rows = list(rows_by_sku.values())
        return self._preview(
            file_name, "legacy_reference_v1", "legacy_reference", "Старый план + Рецептура 50", rows,
            ["Квант замеса и суточная мощность загружены в штуках для сравнения со справочником ПЦ/КЦ.",
             "Состав рецептур учтён количеством компонентов по каждому готовому артикулу.",
             "Эти данные дополняют актуальный справочник и не заменяют скорости и замесы из файла 12.08.2026."],
        )

    def _parse_generic(self, workbook, file_name: str) -> ImportPreview:
        sheet = workbook.active
        raw_headers = [str(value or "").strip() for value in next(sheet.iter_rows(values_only=True))]
        normalized = {self._normalize(name): index for index, name in enumerate(raw_headers)}
        mapping: dict[str, int] = {}
        for field, aliases in self.COLUMN_ALIASES.items():
            for alias in aliases:
                if self._normalize(alias) in normalized:
                    mapping[field] = normalized[self._normalize(alias)]
                    break
        missing = sorted(self.REQUIRED - mapping.keys())
        if missing:
            translated = {"sku": "Артикул", "quantity": "Количество", "requested_date": "Желаемая дата", "due_date": "Крайняя дата"}
            raise ValueError("Не удалось распознать поддерживаемый шаблон; отсутствуют колонки: " + ", ".join(translated[name] for name in missing))

        rows: list[ImportRow] = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in values):
                continue
            errors: list[str] = []
            sku = str(self._value(values, mapping.get("sku")) or "").strip()
            product_name = str(self._value(values, mapping.get("product_name")) or sku).strip()
            quantity = self._decimal(self._value(values, mapping.get("quantity")), errors, "Количество")
            requested = self._date(self._value(values, mapping.get("requested_date")), errors, "Желаемая дата")
            due = self._date(self._value(values, mapping.get("due_date")), errors, "Крайняя дата")
            if not sku:
                errors.append("Не указан артикул")
            if quantity is not None and quantity <= 0:
                errors.append("Количество должно быть больше нуля")
            if requested and due and due < requested:
                errors.append("Крайняя дата раньше желаемой")
            priority_value = self._value(values, mapping.get("priority"))
            try:
                priority = int(priority_value) if priority_value not in (None, "") else 100
            except (ValueError, TypeError):
                priority = 100
                errors.append("Некорректный приоритет")
            rows.append(ImportRow(
                row_number=row_number, sku=sku, product_name=product_name, quantity=quantity,
                source_quantity=quantity, source_unit="кг", quantity_kg=quantity,
                requested_date=requested, due_date=due, priority=priority,
                customer=str(self._value(values, mapping.get("customer")) or "").strip() or None,
                valid=not errors, errors=errors,
            ))
        return self._preview(file_name, "default_v1", "generic", sheet.title, rows, ["Количество трактуется как кг."])

    def _fk_reference(self, workbook) -> dict[str, dict]:
        if "Справочник ФК" not in workbook.sheetnames:
            return {}
        sheet = workbook["Справочник ФК"]
        result: dict[str, dict] = {}
        for values in sheet.iter_rows(min_row=2, values_only=True):
            sku = self._sku(values[0] if values else None)
            if not sku:
                continue
            result[sku] = {
                "line": str(values[2] if len(values) > 2 and values[2] else "").strip() or None,
                "category": str(values[3] if len(values) > 3 and values[3] else "").strip() or None,
                "speed": self._number(values[4] if len(values) > 4 else None),
            }
        return result

    @staticmethod
    def parse_packaging(name: str) -> dict[str, Decimal | None]:
        normalized = name.lower().replace("ё", "е")
        grams = list(re.finditer(r"(\d+(?:[.,]\d+)?)\s*(?:г|гр)(?![а-я])", normalized))
        unit_weight = Decimal(grams[0].group(1).replace(",", ".")) / Decimal("1000") if grams else None
        unit_end = grams[0].end() if grams else 0
        if unit_weight is None:
            kg_unit = re.search(r"(\d+(?:[.,]\d+)?)\s*кг\s*(?=\*|x|х|×)", normalized)
            bare_grams = re.search(r"(\d{2,4}(?:[.,]\d+)?)\s*(?=\*|x|х|×)", normalized)
            if kg_unit:
                unit_weight = Decimal(kg_unit.group(1).replace(",", "."))
                unit_end = kg_unit.end()
            elif bare_grams:
                unit_weight = Decimal(bare_grams.group(1).replace(",", ".")) / Decimal("1000")
                unit_end = bare_grams.end()
        box_weight = None
        box_match = re.search(r"\((\d+(?:[.,]\d+)?)\s*кг", normalized)
        if box_match:
            box_weight = Decimal(box_match.group(1).replace(",", "."))
        elif unit_weight:
            tail = normalized[unit_end:].split("(")[0]
            factors = re.findall(r"(?:\*|x|х|×)\s*(\d+(?:[.,]\d+)?)", tail)
            if factors:
                multiplier = Decimal("1")
                for factor in factors:
                    multiplier *= Decimal(factor.replace(",", "."))
                box_weight = unit_weight * multiplier
        units_per_box = (box_weight / unit_weight).quantize(Decimal("0.001")) if unit_weight and box_weight else None
        return {"unit_weight_kg": unit_weight, "units_per_box": units_per_box, "box_weight_kg": box_weight}

    @staticmethod
    def _preview(file_name: str, mapping: str, template: str, sheet: str, rows: list[ImportRow], notes: list[str]) -> ImportPreview:
        return ImportPreview(
            file_name=file_name, mapping_code=mapping, template_type=template, detected_sheet=sheet,
            total_rows=len(rows), valid_rows=sum(row.valid for row in rows), invalid_rows=sum(not row.valid for row in rows),
            columns=[], rows=rows, notes=notes,
        )

    @staticmethod
    def _year_from_workbook(sheet, file_name: str) -> int:
        for value in (sheet.cell(1, 1).value, sheet.cell(2, 4).value, file_name):
            match = re.search(r"20\d{2}", str(value or ""))
            if match:
                return int(match.group(0))
        return date.today().year

    @staticmethod
    def _sku(value) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _number(value) -> Decimal | None:
        if value in (None, "") or isinstance(value, bool):
            return None
        try:
            number = Decimal(str(value).replace(" ", "").replace(",", "."))
            return number if number.is_finite() else None
        except (InvalidOperation, TypeError, ValueError):
            return None

    @staticmethod
    def _display_number(value: Decimal) -> str:
        rounded = value.quantize(Decimal("0.001"))
        return format(rounded, "f").rstrip("0").rstrip(".")

    @staticmethod
    def _plain_date(value) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        return value if isinstance(value, date) else None

    @staticmethod
    def _header_date(value, year: int) -> date | None:
        parsed = ExcelImportService._plain_date(value)
        if parsed:
            return parsed
        match = re.fullmatch(r"\s*(\d{1,2})\s+([а-яё]+)\s*", str(value or "").lower())
        if not match:
            return None
        months = {
            "янв": 1, "января": 1, "фев": 2, "февраля": 2, "мар": 3, "марта": 3,
            "апр": 4, "апреля": 4, "май": 5, "мая": 5, "июн": 6, "июня": 6,
            "июл": 7, "июля": 7, "авг": 8, "августа": 8, "сен": 9, "сентября": 9,
            "окт": 10, "октября": 10, "ноя": 11, "ноября": 11, "дек": 12, "декабря": 12,
        }
        month = months.get(match.group(2).rstrip("."))
        return date(year, month, int(match.group(1))) if month else None

    @staticmethod
    def _normalize(value: str) -> str:
        return " ".join(value.lower().replace("_", " ").split())

    @staticmethod
    def _value(values: tuple, index: int | None):
        return values[index] if index is not None and index < len(values) else None

    @staticmethod
    def _decimal(value, errors: list[str], label: str) -> Decimal | None:
        try:
            return Decimal(str(value).replace(" ", "").replace(",", "."))
        except (InvalidOperation, TypeError):
            errors.append(f"{label}: ожидается число")
            return None

    @staticmethod
    def _date(value, errors: list[str], label: str) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    pass
        errors.append(f"{label}: ожидается дата")
        return None
